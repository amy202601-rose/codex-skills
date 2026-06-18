#!/usr/bin/env python
import argparse
import datetime as dt
import json
import os
import sys
from pathlib import Path

import openpyxl


DEFAULT_WORKBOOK = r"C:\Users\33663\Desktop\宥宥\01.xlsx"

PLATFORM_COLUMNS = {
    "小红书": "xiaohongshu",
    "抖音": "douyin",
    "视频号": "wechat_channels",
}

REQUIRED_COLUMNS = ["日期", "时间", "视频", "标题", "描述", "标签"]


def parse_date(value):
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = str(value).strip()
    for fmt in ("%Y.%m.%d", "%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Unrecognized date: {value!r}")


def parse_time(value):
    if isinstance(value, dt.datetime):
        return value.time().replace(microsecond=0)
    if isinstance(value, dt.time):
        return value.replace(microsecond=0)
    if isinstance(value, (int, float)):
        seconds = round(float(value) * 24 * 60 * 60)
        return (dt.datetime.min + dt.timedelta(seconds=seconds)).time()
    text = str(value).strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return dt.datetime.strptime(text, fmt).time()
        except ValueError:
            pass
    raise ValueError(f"Unrecognized time: {value!r}")


def should_publish(value):
    text = str(value or "").strip()
    return "发布" in text and "不发布" not in text


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def read_plan(workbook_path, target_date, include_future):
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [cell_text(cell.value) for cell in sheet[1]]
    index = {name: pos for pos, name in enumerate(headers) if name}
    missing = [name for name in REQUIRED_COLUMNS if name not in index]
    missing_platforms = [name for name in PLATFORM_COLUMNS if name not in index]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")
    if missing_platforms:
        raise ValueError(f"Missing platform columns: {', '.join(missing_platforms)}")

    now = dt.datetime.now()
    rows = []
    jobs = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value is not None and str(value).strip() for value in row):
            continue
        row_date = parse_date(row[index["日期"]])
        row_time = parse_time(row[index["时间"]])
        scheduled_at = dt.datetime.combine(row_date, row_time)
        if row_date != target_date:
            continue
        if not include_future and scheduled_at > now:
            continue

        record = {
            "row": row_number,
            "date": row_date.isoformat(),
            "time": row_time.strftime("%H:%M:%S"),
            "scheduled_at": scheduled_at.isoformat(timespec="seconds"),
            "video": cell_text(row[index["视频"]]),
            "cover": cell_text(row[index["封面"]]) if "封面" in index else "",
            "title": cell_text(row[index["标题"]]),
            "description": cell_text(row[index["描述"]]),
            "tags": cell_text(row[index["标签"]]),
            "platforms": [],
            "exists": {},
        }
        record["exists"]["video"] = bool(record["video"] and os.path.exists(record["video"]))
        record["exists"]["cover"] = bool(record["cover"] and os.path.exists(record["cover"]))
        for column_name, platform_id in PLATFORM_COLUMNS.items():
            flag = row[index[column_name]]
            if should_publish(flag):
                record["platforms"].append(platform_id)
                jobs.append(
                    {
                        "row": row_number,
                        "platform": platform_id,
                        "platform_label": column_name,
                        "scheduled_at": record["scheduled_at"],
                        "video": record["video"],
                        "cover": record["cover"],
                        "title": record["title"],
                        "description": record["description"],
                        "tags": record["tags"],
                        "exists": record["exists"],
                    }
                )
        rows.append(record)

    return {
        "workbook": str(Path(workbook_path)),
        "sheet": sheet.title,
        "target_date": target_date.isoformat(),
        "include_future": include_future,
        "rows": rows,
        "jobs": jobs,
    }


def main():
    parser = argparse.ArgumentParser(description="Read an Excel video publishing schedule.")
    parser.add_argument("--workbook", default=DEFAULT_WORKBOOK)
    parser.add_argument("--date", default=dt.date.today().isoformat(), help="Target date, YYYY-MM-DD.")
    parser.add_argument("--include-future", action="store_true", help="Include today's future scheduled rows.")
    args = parser.parse_args()

    try:
        target_date = dt.date.fromisoformat(args.date)
        plan = read_plan(args.workbook, target_date, args.include_future)
    except Exception as exc:
        print(json.dumps({"error": str(exc)}, ensure_ascii=False), file=sys.stderr)
        return 1

    print(json.dumps(plan, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
